from typing import List

from openpyxl import Workbook, load_workbook
from datetime import datetime
import os
import subprocess
import numpy as np

# checks if epoch number == (1, 2, 5, 10, 20, 50...)
def Display_Check(Epoch):
    if Epoch % 5000 == 0:
        return True
    while Epoch % 10 == 0:
        Epoch //= 10
        if (Epoch == 0):
            break
        
    if (Epoch in (1, 2, 5)):
        return True
    
    return False

def Write_To_TempData(numPoints: int, EpochNum: int, matError: np.ndarray):
    file_path_temp = "Temp_Log.xlsx"
    file_path_txt = "Temp_Holder.txt"
    
    wb = try_temp_workbook(file_path_temp)
    sheet = wb["TEMP_DATA"]

    # formatting
    if EpochNum == 1:
        header = ["Epoch#"] + [f"E{i + 1}" for i in range(numPoints)]
        sheet.append(header)
    errorList: list = matError.flatten().tolist()
    sheet.append([EpochNum] + errorList)

    # adds the data
    """
    with open(file_path_txt, "r") as txt:
        for line in txt:
            line = line.strip()
            columns = [col.strip() for col in line.split(",")]
            
            # convert numbers to float if necessary
            for i in range(1, len(columns)):
                columns[i] = float(columns[i])
            sheet.append(columns)
        """

    """
    Stuff for Epoch holder not needed here
    
    Epoch_Data_Line = Calculate_Epoch_Data((EpochNum), errorList)

    sheet = wb["EPOCH_HOLDER"]

    # formatting
    if EpochNum == 1:
        sheet.append([f"Epoch #", "Min Error", "Quartile 1", "Median Error", "Quartile 3", "Max Error", "IQR", "Error Range", "Mean Error", "STDV"])

    sheet.append(Epoch_Data_Line)

    wb.save(file_path_temp)
    """

def Write_To_TempEpoch(EpochNum: int, matError: np.ndarray):
    file_path_temp = "Temp_Log.xlsx"
    wb = try_temp_workbook(file_path_temp)
    sheet = wb["EPOCH_HOLDER"]

    if EpochNum == 1:
        sheet.append(
            [f"Epoch #", "Min Error", "Quartile 1", "Median Error", "Quartile 3", "Max Error", "IQR", "Error Range",
             "Mean Error", "STDV"])

    errorList: list = matError.flatten().tolist()
    Epoch_Data_Line = Calculate_Epoch_Data((EpochNum), errorList)

    sheet.append(Epoch_Data_Line)
    wb.save(file_path_temp)

def Write_To_xlsx():
    file_main = "ML_log.xlsx"
    file_temp = "Temp_Log.xlsx"
    file_stage = file_main + ".tmp"  # short-lived safety file

    # Load temp workbook (scratch file — allowed to break)
    wb_temp = load_workbook(file_temp)
    sheet_temp = wb_temp["EPOCH_HOLDER"]

    # Load main workbook (must be protected)
    wb_main = load_workbook(file_main)
    sheet_main = wb_main["Epoch Data"]

    # Append rows (skip header row from temp)
    for i, row in enumerate(sheet_temp.iter_rows(values_only=True)):
        #if i == 0:
            #continue
        sheet_main.append(row)

    # Save safely to staging file first
    wb_main.save(file_stage)

    wb_temp.close()
    wb_main.close()

    # Atomic replace protects main file
    os.replace(file_stage, file_main)

"""
def Write_To_txt(matWeight: np.ndarray, matBias: np.ndarray, Quality: np.ndarray, Error: np.ndarray) -> None:
    file_path = "Temp_Holder.txt"
    
    with open(file_path, "a") as f:
        weights_str = ", ".join(str(w) for w in matWeight)
        prediction = Quality - Error
        
        f.write(
            f"{weights_str}, {matBias}, {Quality}, {prediction}, {Error}\n"
        ) 
"""


def Calculate_Epoch_Data(EpochNum: int, ErrorList: list) -> list:
    errors = np.abs(np.array(ErrorList))

    min_err = np.min(errors)
    q1 = np.percentile(errors, 25)
    median = np.percentile(errors, 50)
    q3 = np.percentile(errors, 75)
    max_err = np.max(errors)

    iqr = q3 - q1
    err_range = max_err - min_err
    mean = np.mean(errors)
    stdv = np.std(errors)

    return [
        EpochNum,
        min_err,
        q1,
        median,
        q3,
        max_err,
        iqr,
        err_range,
        mean,
        stdv
    ]
def Log_Tests(testErrors):
    file_path = "Temp_Log.xlsx"

    wb = load_workbook(file_path)
    sheet = wb["EPOCH_HOLDER"]

    sheet.append(testErrors)
    wb.save(file_path)

def Open_xlsm(file_path):
    subprocess.call(["open", file_path])


def Save_Close(file_path):
    """Save and close an Excel workbook on macOS using AppleScript."""
    file_name = os.path.basename(file_path)

    script = f'''
    tell application "Microsoft Excel"
        if it is running then
            if (name of workbooks) contains "{file_name}" then
                save workbook "{file_name}"
                close workbook "{file_name}"
            end if
        end if
    end tell
    '''

    subprocess.run(["osascript", "-e", script], capture_output=True, text=True)

def try_temp_workbook(file_path):
    try:
        wb = load_workbook(file_path)
    except Exception:
        # File missing or corrupted → recreate
        if os.path.exists(file_path):
            os.remove(file_path)

        wb = Workbook()
        wb.remove(wb.active)  # remove default sheet

        wb.create_sheet("TEMP_DATA")
        wb.create_sheet("EPOCH_HOLDER")

        wb.save(file_path)

    # Ensure required sheets exist
    required_sheets = ["TEMP_DATA", "EPOCH_HOLDER"]
    for name in required_sheets:
        if name not in wb.sheetnames:
            wb.create_sheet(name)

    return wb

def Clear_Temp(file_path: str) -> None:
    try:
        os.remove(file_path)
    except FileNotFoundError:
        pass