#!/usr/bin/env python3
"""Initialize ML_log.xlsx with Temp Dump sheet for logging."""

import os
from openpyxl import Workbook

def init_log():
    file_path = 'ML_log.xlsx'
    if not os.path.exists(file_path):
        wb = Workbook()
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        wb.create_sheet('Temp Dump')
        wb.save(file_path)
        print(f'Created {file_path} with Temp Dump sheet')
    else:
        print(f'{file_path} already exists')

if __name__ == '__main__':
    init_log()
